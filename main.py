import os

import openpyxl as op
from dotenv import load_dotenv
from loguru import logger

from utils.chat_gigachat import get_gigachat_response
from utils.data_reader import open_list_gup
from utils.file_operations import create_folder, create_md_file

# Загрузка переменных окружения
load_dotenv(dotenv_path='settings/config.env')
GIGA_CHAT = os.getenv('GIGA_CHAT')

# Настройка логирования
logger.add("logs/app.log", format="{time} {level} {message}", level="INFO", rotation="1 MB")


def filling_data_md_file():
    """
    Функция заполнения данных в md файл.

     Основная функция программы.
    Получает данные из Excel, отправляет запросы в GigaChat
    и сохраняет результаты в Markdown файлы.
    """
    # Получаем данные из Excel
    list_gup = open_list_gup('file.xlsx')

    # Обработка каждого участка и профессии
    for plot, work in list_gup:
        logger.info(f"Обработка участка: {plot}, профессии: {work}")

        # Создаем папку для участка
        folder_path = f'test/{plot}'
        create_folder(folder_path)

        # Формируем запрос к GigaChat
        system_prompt = f"Расскажите, чем занимается данная профессия на предприятии: Участок: {plot}, Профессия: {work}"
        response = get_gigachat_response(GIGA_CHAT, system_prompt)

        # Сохраняем ответ в файл
        file_path = f'{folder_path}/{work}'
        create_md_file(file_path, response)


def PrOMT_for_artificial_intelligence(plot):
    """
    Функция заполнения данных в md файл.

    Получает данные из Excel, отправляет запросы в GigaChat
    и сохраняет результаты в Markdown файлы.
    """
    system_prompt = (f"Я заполняю данные для аттестации рабочих мест, и мне нужны краткое описание работы (протокол "
                     f"напряженности), которое касается данной профессии. Ответ нужен краткий, на 10 слов максимум: "
                     f"Профессия: {plot}")
    return system_prompt


def Opening_Excel_file(parsing_file):
    """
    Функция заполнения данных в Excel файл.

    Получает данные из Excel, отправляет запросы в GigaChat
    и сохраняет результаты в Excel файл.
    :return: None
    """
    wb = op.load_workbook(parsing_file)
    ws = wb.active
    return wb, ws


def Record_in_Excel_File(parsing_file, ws, wb, row_idx, response, column_write):
    """
    Функция заполнения данных в Excel файл.

    Получает данные из Excel, отправляет запросы в GigaChat
    и сохраняет результаты в Excel файл.
    :return: None
    """
    # Записываем ответ в 3-ю колонку
    ws.cell(row=row_idx, column=column_write, value=response)
    # Сохраняем изменения в файле
    wb.save(parsing_file)


def Opening_Excel_Parsing_Processing_through_Article_Intercue_Record_in_Excel_File(parsing_file, column_write):
    """
    Функция заполнения данных в Excel файл.

    Получает данные из Excel, отправляет запросы в GigaChat
    и сохраняет результаты в Excel файл.
    :param parsing_file: путь к Excel файлу для парсинга и заполнения данных.
    :return: None
    """
    wb, ws = Opening_Excel_file(parsing_file)
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=116, min_col=1, max_col=1), start=5):
        plot = [cell.value for cell in row]
        if plot[0] is None:
            continue
        else:
            # Формируем запрос к GigaChat
            system_prompt = PrOMT_for_artificial_intelligence(plot)
            response = get_gigachat_response(GIGA_CHAT, system_prompt)
            print(f"Профессия: {plot}. Ответ: {response}")
            Record_in_Excel_File(parsing_file, ws, wb, row_idx, response, column_write)


def main():
    """
    Основное меню программы.
    """
    logger.info("Запуск программы")
    try:
        print(
            "1 - Заполнение данных в md файл\n"
            "2 - Заполнение данных в Excel (Оборудование)\n"
            "3 - Заполнение данных в Excel (Сырье)\n"
            "4 - Заполнение данных в Excel (Сведения об источнике вредного фактора (для перечня РМ))\n"
            "5 - Заполнение данных в Excel (Краткое описание работы (протокол тяжести))\n"
            "6 - Заполнение данных в Excel (Краткое описание работы (протокол напряженности))\n"
        )
        user_input = input("Введите номер:")
        if user_input == '1':  # Заполнение данных в md файл
            filling_data_md_file()
        elif user_input == '2':  # Заполнение данных в Excel
            wb = op.load_workbook('data.xlsx')
            ws = wb.active
            for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=116, min_col=1, max_col=1), start=5):
                plot = [cell.value for cell in row]
                if plot[0] is None:
                    continue
                else:
                    # Формируем запрос к GigaChat
                    system_prompt = f"Я заполняю данные для аттестации рабочих мест и мне нужно Оборудование, которое использует данная профессия. Ответ нужен краткий, на 8 слов максимум: Профессия: {plot}"
                    response = get_gigachat_response(GIGA_CHAT, system_prompt)
                    print(f"Профессия: {plot}. Ответ: {response}")
                    # Записываем ответ в 3-ю колонку
                    ws.cell(row=row_idx, column=3, value=response)
                    # Сохраняем изменения в файле
                    wb.save('data.xlsx')
        elif user_input == '3':  # Заполнение данных в Excel
            wb = op.load_workbook('data.xlsx')
            ws = wb.active
            for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=116, min_col=1, max_col=1), start=5):
                plot = [cell.value for cell in row]
                if plot[0] is None:
                    continue
                else:
                    # Формируем запрос к GigaChat
                    system_prompt = f"Я заполняю данные для аттестации рабочих мест и мне нужно Материалы и сырье, которое использует данная профессия. Ответ нужен краткий, на 8 слов максимум: Профессия: {plot[0]}"
                    response = get_gigachat_response(GIGA_CHAT, system_prompt)
                    print(f"Профессия: {plot}. Ответ: {response}")
                    # Записываем ответ в 4-ю колонку
                    ws.cell(row=row_idx, column=4, value=response)
                    # Сохраняем изменения в файле
                    wb.save('data.xlsx')
        elif user_input == '4':  # Заполнение данных в Excel
            wb = op.load_workbook('data.xlsx')
            ws = wb.active
            for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=116, min_col=1, max_col=1), start=5):
                plot = [cell.value for cell in row]
                if plot[0] is None:
                    continue
                else:
                    # Формируем запрос к GigaChat
                    system_prompt = f"Я заполняю данные для аттестации рабочих мест, и мне нужны сведения об источнике вредного фактора (для перечня РМ), которое касается данной профессии. Ответ нужен краткий, на 10 слов максимум: Профессия: {plot}"
                    response = get_gigachat_response(GIGA_CHAT, system_prompt)
                    print(f"Профессия: {plot}. Ответ: {response}")
                    # Записываем ответ в 3-ю колонку
                    ws.cell(row=row_idx, column=3, value=response)
                    # Сохраняем изменения в файле
                    wb.save('data.xlsx')
        elif user_input == '5':  # Заполнение данных в Excel
            wb = op.load_workbook('data.xlsx')
            ws = wb.active
            for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=116, min_col=1, max_col=1), start=5):
                plot = [cell.value for cell in row]
                if plot[0] is None:
                    continue
                else:
                    # Формируем запрос к GigaChat
                    system_prompt = f"Я заполняю данные для аттестации рабочих мест, и мне нужны краткое описание работы (протокол тяжести), которое касается данной профессии. Ответ нужен краткий, на 10 слов максимум: Профессия: {plot}"
                    response = get_gigachat_response(GIGA_CHAT, system_prompt)
                    print(f"Профессия: {plot}. Ответ: {response}")

        elif user_input == '6':  # Заполнение данных в Excel
            parsing_file = "data.xlsx"
            Opening_Excel_Parsing_Processing_through_Article_Intercue_Record_in_Excel_File(parsing_file=parsing_file, column_write=5)
        else:
            print("Неверный ввод")

    except Exception as e:
        logger.exception(f"Ошибка в программе: {e}")


if __name__ == "__main__":
    main()
